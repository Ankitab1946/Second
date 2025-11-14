# app.py - Final auto-detect header engine (R1), A3 formatting, notes in Column A
import streamlit as st
import pandas as pd
import re
import os
from io import BytesIO

# optional imports (try/except to avoid hard crash if not installed)
try:
    from openpyxl import load_workbook
    import openpyxl
except Exception:
    load_workbook = None
try:
    import pyxlsb
except Exception:
    pyxlsb = None

st.set_page_config(page_title="Financial Extractor (Auto Header Detect, R1)", layout="wide")

# -----------------------
# Config
# -----------------------
IGNORED_STATUS_WORDS = {
    "restated", "provisional", "unaudited", "reclassified",
    "notes", "revised", "converted", "normalized", "n.a.", "n.a", "na", "unaudited/unauthorised"
}

VALID_PARENT_KEYWORDS = {
    "historical annual", "historical annuals", "historical interims",
    "historical", "annual", "interim", "forecasts", "forecast",
    "initial budget", "budget", "variation", "cagrars", "cagr"
}

# A3 formatter: TitleCase and underscores
def format_token_for_output(token: str) -> str:
    if not token:
        return ""
    t = " ".join(str(token).split())  # normalize whitespace
    t = t.lower()
    # remove trailing punctuation
    t = re.sub(r"^[\s\W]+|[\s\W]+$", "", t)
    words = [w.capitalize() for w in re.split(r"\s+", t) if w]
    return "_".join(words)

# token validator (strict rules)
def is_valid_header_token(tok: str) -> bool:
    if not tok:
        return False
    s = str(tok).strip()
    if not s:
        return False
    t = s.lower()
    if t in IGNORED_STATUS_WORDS:
        return False
    # 4-digit year
    if re.fullmatch(r"\d{4}", t):
        return True
    # year range like 2020-2024 / 2020_2024 / 2020–2024
    if re.fullmatch(r"\d{4}[\-_–]\d{4}", t):
        return True
    # periods like 1H, 2H, Q1-Q4, LTM
    if re.fullmatch(r"[12]h", t) or re.fullmatch(r"q[1-4]", t) or t == "ltm":
        return True
    # parent detection (fuzzy)
    base = re.sub(r"[^\w\s]", " ", t).replace("  ", " ").strip()
    if "historical" in base and "annual" in base:
        return True
    if "historical" in base and "interim" in base:
        return True
    for kw in VALID_PARENT_KEYWORDS:
        if kw in base:
            return True
    # otherwise reject numeric garbage, decimals, floats
    if re.fullmatch(r"-?\d+\.\d+", t):
        return False
    # reject plain numbers that are not years
    if t.isdigit():
        return False
    return False

# clean values: map status words to NA, convert numbers, handle percent and parentheses
def clean_value(x):
    if pd.isna(x):
        return pd.NA
    s = str(x).strip()
    if not s:
        return pd.NA
    ls = s.lower()
    if ls in IGNORED_STATUS_WORDS:
        return pd.NA
    # percent
    if s.endswith("%"):
        try:
            return float(s[:-1].replace(",", "").strip()) / 100.0
        except Exception:
            return s
    # parentheses negative
    if re.fullmatch(r"\(\s*[\d,\.]+\s*\)", s):
        try:
            return -float(s.strip("()").replace(",", ""))
        except:
            return s
    # numeric
    s_clean = s.replace(",", "")
    if re.fullmatch(r"-?\d+(\.\d+)?", s_clean):
        try:
            return float(s_clean)
        except:
            return s
    return s

# dedupe column names to avoid duplicates
def dedupe_columns(cols):
    out, counts = [], {}
    for c in cols:
        k = "" if c is None else str(c)
        if k not in counts:
            counts[k] = 0
            out.append(k)
        else:
            counts[k] += 1
            out.append(f"{k}_{counts[k]}")
    return out

# produce excel bytes for download
def to_excel(groups):
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for cat, dfs in groups.items():
            for i, df in enumerate(dfs, 1):
                safe_sheet = (cat[:24] + f"_{i}") if cat else f"Sheet_{i}"
                try:
                    df.to_excel(writer, index=False, sheet_name=safe_sheet)
                except Exception:
                    # fallback name
                    df.to_excel(writer, index=False, sheet_name=f"Sheet{i}")
    out.seek(0)
    return out

# -----------------------
# Helpers to read excel raw with header=None
# -----------------------
def read_excel_sheets_openpyxl(path):
    # use pandas.ExcelFile with openpyxl engine to parse into dataframes header=None
    try:
        xlf = pd.ExcelFile(path, engine="openpyxl")
    except Exception as e:
        raise
    sheets = {}
    for s in xlf.sheet_names:
        try:
            sheets[s] = xlf.parse(s, header=None, dtype=object)
        except Exception:
            sheets[s] = None
    return sheets

def read_xls_sheets(path):
    xlf = pd.ExcelFile(path, engine="xlrd")
    sheets = {}
    for s in xlf.sheet_names:
        try:
            sheets[s] = xlf.parse(s, header=None, dtype=object)
        except Exception:
            sheets[s] = None
    return sheets

def read_xlsb_sheets(path):
    sheets = {}
    try:
        xlf = pd.ExcelFile(path, engine="pyxlsb")
    except Exception:
        return {}
    for s in xlf.sheet_names:
        try:
            sheets[s] = pd.read_excel(path, sheet_name=s, header=None, engine="pyxlsb", dtype=object)
        except Exception:
            sheets[s] = None
    return sheets

def read_csv_sheet(path):
    return {"csv": pd.read_csv(path, header=None, dtype=object)}

# pdf extraction (best-effort)
def extract_tables_from_pdf(path):
    out = []
    # camelot lattice
    try:
        import camelot
        ct = camelot.read_pdf(path, pages="all", flavor="lattice")
        for t in ct:
            out.append(t.df)
    except Exception:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            for p in pdf.pages:
                for tbl in p.extract_tables():
                    out.append(pd.DataFrame(tbl))
    except Exception:
        pass
    try:
        import tabula
        tbs = tabula.read_pdf(path, pages="all", multiple_tables=True)
        for t in tbs:
            out.append(t)
    except Exception:
        pass
    return out

# -----------------------
# Auto-detect header band & build headers
# -----------------------
def detect_header_band_and_build(df_raw, sheet_ws=None, debug=False):
    """
    df_raw: pandas DataFrame header=None representing entire sheet (rows correspond to excel rows starting at 0)
    sheet_ws: openpyxl worksheet object (optional, for merged detection and column widths)
    Returns:
        master_notes_str, headers_list, data_df (data rows only, header rows removed)
    """
    # drop fully empty rows for scanning but keep mapping to original indexes
    # convert full df to string for scanning
    nrows, ncols = df_raw.shape
    # build a small helper to get cell value (string) safely
    def cell_str(r, c):
        try:
            v = df_raw.iat[r, c]
            return "" if pd.isna(v) else str(v).strip()
        except Exception:
            return ""

    # find candidate period row: row with many 4-digit years or period tokens
    def row_period_score(r):
        score = 0
        for c in range(ncols):
            v = cell_str(r, c)
            if re.fullmatch(r"\d{4}", v):
                score += 2
            elif re.fullmatch(r"\d{4}[\-_–]\d{4}", v):
                score += 2
            elif re.fullmatch(r"[12]H", v, flags=re.IGNORECASE) or re.fullmatch(r"Q[1-4]", v, flags=re.IGNORECASE) or v.strip().lower() == "ltm":
                score += 1
        return score

    best_row = None
    best_score = -1
    # scan all rows for period candidates
    for r in range(min(30, nrows)):  # usually header bands appear in first ~30 rows
        s = row_period_score(r)
        if debug:
            pass
        if s > best_score and s > 0:
            best_score = s
            best_row = r

    # If not found in first 30, scan entire sheet (fallback)
    if best_row is None:
        for r in range(nrows):
            s = row_period_score(r)
            if s > best_score and s > 0:
                best_score = s
                best_row = r

    if best_row is None:
        # fallback: try to find a row with multiple numeric like 2020-2025 or '2020' occurrences less strictly
        for r in range(min(50, nrows)):
            cnt_years = sum(1 for c in range(ncols) if re.search(r"\b20\d{2}\b", cell_str(r, c)))
            if cnt_years >= 2:
                best_row = r
                break

    # Determine parent row: scan upwards from period row to find a row containing parent keyword
    parent_row = None
    meta_row = None
    if best_row is not None:
        # look up to 6 rows above for parent indicator
        for up in range(1, 7):
            rr = best_row - up
            if rr < 0:
                break
            # combine row text
            row_text = " ".join([cell_str(rr, c) for c in range(ncols)])
            if row_text.strip():
                low = row_text.strip().lower()
                # if row contains parent keywords, choose it
                if any(kw in low for kw in VALID_PARENT_KEYWORDS) or ("historical" in low and "annual" in low) or ("historical" in low and "interim" in low):
                    parent_row = rr
                    break
        # if not found, choose nearest non-empty above period row (but not too far)
        if parent_row is None:
            for up in range(1, 8):
                rr = best_row - up
                if rr < 0:
                    break
                row_text = " ".join([cell_str(rr, c) for c in range(ncols)])
                if row_text.strip():
                    parent_row = rr
                    break

        # meta row: often right below period row (e.g., "Restated"), so check best_row+1
        below = best_row + 1
        if below < nrows:
            row_text = " ".join([cell_str(below, c) for c in range(ncols)])
            if any(s.lower() in row_text.lower() for s in IGNORED_STATUS_WORDS):
                meta_row = below

    # If still no parent_row or period row, fallback: assume header at top rows 0..2
    if best_row is None:
        parent_row = 0
        best_row = 1
        meta_row = 2 if nrows > 2 else None

    # Now build headers using parent_row (P), period_row (R), meta_row (M)
    P = parent_row
    R = best_row
    M = meta_row

    # Build master notes: all rows above P (0..P-1)
    notes_rows = []
    for r in range(0, P):
        joined = " ".join([cell_str(r, c) for c in range(ncols)]).strip()
        if joined:
            notes_rows.append(joined)
    master_notes = " | ".join(notes_rows).strip()
    # Format master notes header for Column A (A3 style)
    master_notes_header = format_token_for_output(master_notes) if master_notes else "Notes"

    # Build header tokens for each column: Parent + (maybe period row token) + (maybe meta if meaningful)
    headers = []
    for c in range(ncols):
        parts = []
        # Parent token: prefer merged cells if sheet_ws provided
        ptoken = ""
        if sheet_ws is not None:
            try:
                # openpyxl uses 1-based indexing
                raw = sheet_ws.cell(row=P+1, column=c+1).value
                if raw is not None and str(raw).strip():
                    ptoken = str(raw).strip()
                else:
                    # maybe merged parent exists above (see merged ranges)
                    ptoken = ""
            except Exception:
                ptoken = ""

        if not ptoken:
            ptoken = cell_str(P, c)

        # include parent token only if valid-ish and not a pure number
        if ptoken and is_valid_header_token(ptoken):
            parts.append(ptoken)

        # sub/period token from period row R
        rtoken = cell_str(R, c)
        if rtoken and is_valid_header_token(rtoken):
            parts.append(rtoken)

        # meta (ignored per R1) -> do not include status words; but include if meta seems like "Variation" which is a header
        mtoken = ""
        if M is not None:
            mtoken = cell_str(M, c)
            if mtoken:
                lowm = mtoken.strip().lower()
                # include if it's a meaningful header token like 'variation' or 'ltm' or 'variation' etc.
                if lowm in ("variation", "variation%", "variance") or is_valid_header_token(mtoken):
                    # but do not include Restated or similar (we ignore)
                    if lowm not in IGNORED_STATUS_WORDS:
                        parts.append(mtoken)

        # If parts empty, but period row has year (even if parent missing) - keep year (we will try to salvage)
        if not parts:
            if rtoken and re.fullmatch(r"\d{4}", rtoken.strip()):
                parts.append(rtoken.strip())
        # If still empty -> empty header
        if parts:
            # Format parts into A3 tokens
            fmt_parts = [format_token_for_output(p) for p in parts]
            headers.append("_".join(fmt_parts))
        else:
            headers.append("")  # will be dropped later

    # Now build data frame removing header rows (rows 0.. up to R, plus possible meta row)
    drop_upto = R
    # We want data starting from first data row which is R+1 if meta is R+1 then R+2 etc.
    start_row = R + 1
    if M is not None and M == R + 1:
        start_row = M + 1

    # Ensure start_row within bounds
    if start_row >= nrows:
        start_row = min(nrows-1, R+1)

    data_df = df_raw.iloc[start_row:].reset_index(drop=True).copy()
    # assign headers (we'll set first column name as master_notes_header)
    # but remove columns where headers are empty or deemed comment columns
    # first set temporary columns
    temp_cols = headers.copy()
    # If first col header is empty but df has first col used as particulars, we will set it to 'Particulars' then rename to master header later.
    if temp_cols and (temp_cols[0] == "" or temp_cols[0].lower() in ("nan", "none")):
        temp_cols[0] = "Particulars"
    # apply temp columns
    data_df.columns = temp_cols

    # clean cell values
    data_df = data_df.applymap(clean_value)

    # drop columns where header empty AND column largely empty
    cols_to_keep = []
    for col in data_df.columns:
        if col and col.strip():
            cols_to_keep.append(col)
        else:
            # if column has any non-NA values maybe it's particulars column - keep if so
            non_na = data_df[col].notna().sum()
            if non_na > 0:
                # keep
                idx = col
                cols_to_keep.append(col)
            # else drop

    data_df = data_df.loc[:, cols_to_keep].copy()

    # Now drop comment/noise columns heuristically:
    # For each column except 'Particulars' keep if numeric ratio > 0.3 or header is meaningful
    def is_comment_col(scol):
        series = data_df[scol]
        total = len(series)
        if total == 0:
            return True
        non_numeric = series.apply(lambda v: not (isinstance(v, (int, float)) and pd.notna(v))).sum()
        # if more than 75% non-numeric and header not meaningful, treat as comment
        if non_numeric / total >= 0.75:
            # but if header looks like a valid header token, don't mark as comment
            if is_valid_header_token(scol):
                return False
            # if header is Particulars or master notes placeholder, don't mark
            if scol.lower() in ("particulars", format_token_for_output(master_notes).lower(), "notes"):
                return False
            return True
        return False

    to_drop = [c for c in data_df.columns if is_comment_col(c)]
    for c in to_drop:
        try:
            data_df.drop(columns=[c], inplace=True)
        except Exception:
            pass

    # ensure first column is particulars; rename first column to master_notes_header (A3 formatted)
    cols_final = list(data_df.columns)
    if len(cols_final) == 0:
        return master_notes_header, [], pd.DataFrame()
    # if first column not "Particulars" try find a likely particulars column by searching for strings like 'Revenue', 'Profit', '%' etc in the top rows
    first_col = cols_final[0]
    # rename first column to master_notes header (as user requested Column A header)
    new_cols = data_df.columns.tolist()
    new_cols[0] = master_notes_header
    data_df.columns = new_cols

    # final cleanup: drop all-empty columns
    data_df = data_df.dropna(axis=1, how="all")
    # dedupe column names
    data_df.columns = dedupe_columns(data_df.columns)

    return master_notes_header, headers, data_df

# -----------------------
# Streamlit UI
# -----------------------
st.title("Financial Extractor — Auto Header Detect (R1)")

debug = st.sidebar.checkbox("Enable Debug Mode", False)
uploaded = st.file_uploader("Upload file (.xlsx/.xlsm/.xls/.xlsb/.csv/.pdf)", type=["xlsx", "xlsm", "xls", "xlsb", "csv", "pdf"])
if not uploaded:
    st.stop()

# save file to temp
tmp_path = f"tmp_{uploaded.name}"
with open(tmp_path, "wb") as f:
    f.write(uploaded.getbuffer())

ext = os.path.splitext(tmp_path)[1].lower()

tables = []

if ext in (".xlsx", ".xlsm"):
    if load_workbook is None:
        st.error("openpyxl required for xlsx/xlsm. Install with pip install openpyxl")
        st.stop()
    # use pandas to read raw and openpyxl to access ws for merged ranges
    sheets = pd.ExcelFile(tmp_path, engine="openpyxl").sheet_names
    selected = st.multiselect("Select sheets to extract", sheets, sheets)
    for s in selected:
        try:
            df_raw = pd.read_excel(tmp_path, sheet_name=s, header=None, dtype=object, engine="openpyxl")
            wb = load_workbook(tmp_path, data_only=True)
            ws = wb[s]
            master_notes_header, ps_headers, data_df = detect_header_band_and_build(df_raw, sheet_ws=ws, debug=debug)
            if data_df.empty:
                continue
            tables.append(data_df)
            if debug:
                st.write("Sheet:", s, "master notes header:", master_notes_header)
                st.write("Raw detected headers (raw tokens):", ps_headers[:40])
        except Exception as e:
            st.error(f"Failed to parse sheet {s}: {e}")

elif ext == ".xls":
    sheets = pd.ExcelFile(tmp_path, engine="xlrd").sheet_names
    selected = st.multiselect("Select sheets to extract", sheets, sheets)
    for s in selected:
        try:
            df_raw = pd.read_excel(tmp_path, sheet_name=s, header=None, dtype=object, engine="xlrd")
            # no openpyxl ws available
            master_notes_header, ps_headers, data_df = detect_header_band_and_build(df_raw, sheet_ws=None, debug=debug)
            if data_df.empty:
                continue
            tables.append(data_df)
            if debug:
                st.write("Sheet:", s, "master notes header:", master_notes_header)
                st.write("Raw detected headers:", ps_headers[:40])
        except Exception as e:
            st.error(f"Failed to parse sheet {s}: {e}")

elif ext == ".xlsb":
    if pyxlsb is None:
        st.error("pyxlsb required to read xlsb. pip install pyxlsb")
        st.stop()
    try:
        xlsb_map = read_xlsb_sheets(tmp_path)
        sheets = list(xlsb_map.keys())
        selected = st.multiselect("Select sheets to extract", sheets, sheets)
        for s in selected:
            df_raw = xlsb_map.get(s)
            if df_raw is None:
                continue
            master_notes_header, ps_headers, data_df = detect_header_band_and_build(df_raw, sheet_ws=None, debug=debug)
            if data_df.empty:
                continue
            tables.append(data_df)
            if debug:
                st.write("Sheet:", s, "master notes header:", master_notes_header)
                st.write("Raw detected headers:", ps_headers[:40])
    except Exception as e:
        st.error(f"Failed to read xlsb: {e}")

elif ext == ".csv":
    try:
        df_raw = pd.read_csv(tmp_path, header=None, dtype=object)
        master_notes_header, ps_headers, data_df = detect_header_band_and_build(df_raw, sheet_ws=None, debug=debug)
        if not data_df.empty:
            tables.append(data_df)
        if debug:
            st.write("CSV master notes header:", master_notes_header)
            st.write("Detected headers:", ps_headers[:40])
    except Exception as e:
        st.error(f"Failed to read csv: {e}")

elif ext == ".pdf":
    pdf_tables = extract_tables_from_pdf(tmp_path)
    if not pdf_tables:
        st.warning("No tables extracted from PDF.")
    for i, tdf in enumerate(pdf_tables, 1):
        try:
            # attempt same detection on each table
            master_notes_header, ps_headers, data_df = detect_header_band_and_build(tdf, sheet_ws=None, debug=debug)
            if not data_df.empty:
                tables.append(data_df)
            if debug:
                st.write(f"PDF table {i} master notes header:", master_notes_header)
                st.write("Detected headers:", ps_headers[:40])
        except Exception:
            continue

# Show results
tables = [t for t in tables if not t.empty]
st.success(f"Extracted {len(tables)} table(s)")

groups = {"Balance Sheet": [], "Income Statement": [], "Cash Flow Statement": [], "Other": []}
for idx, df in enumerate(tables, 1):
    st.subheader(f"Extracted Table {idx}")
    st.dataframe(df, width="stretch")
    # classify by column names
    cols_join = " ".join(df.columns).lower()
    if any(k in cols_join for k in ["asset", "balance", "liabil", "equity"]):
        groups["Balance Sheet"].append(df)
    elif any(k in cols_join for k in ["income", "revenue", "profit", "loss"]):
        groups["Income Statement"].append(df)
    elif any(k in cols_join for k in ["cash", "oper", "invest", "financ"]):
        groups["Cash Flow Statement"].append(df)
    else:
        groups["Other"].append(df)

st.header("Summary")
for k, v in groups.items():
    st.write(f"**{k}**: {len(v)} tables")

if tables:
    st.download_button("📥 Download Extracted Financials", data=to_excel(groups), file_name="Extracted_Financials.xlsx")

# cleanup temp files (optional)
try:
    os.remove(tmp_path)
except Exception:
    pass