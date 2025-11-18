# import streamlit as st
# import openpyxl
# import pandas as pd
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
# import io

# # Function to check if a cell has light blue background (adjust RGB if needed)
# def is_light_blue(cell):
#     try:
#         return cell.fill.fgColor.rgb == 'FFD3D3D3'
#     except:
#         return False

# # Function to get merged range for a cell
# def get_merged_range(sheet, cell):
#     for merged_range in sheet.merged_cells.ranges:
#         if cell.coordinate in merged_range:
#             return merged_range
#     return None

# # Function to check if a value is truly blank (None, empty string, or whitespace only)
# def is_blank_value(value):
#     if value is None:
#         return True
#     if isinstance(value, str) and value.strip() == "":
#         return True
#     return False

# # Function to count leading spaces in a string
# def count_leading_spaces(text):
#     if not isinstance(text, str):
#         return 0
#     return len(text) - len(text.lstrip(' '))

# # Function to check if rest of row has values (excluding "Restated")
# def has_row_values(sheet, row, max_col):
#     for col in range(2, max_col + 1):  # Start from B (column 2)
#         col_letter = get_column_letter(col)
#         cell_value = sheet[f'{col_letter}{row}'].value
#         # Skip truly blank
#         if is_blank_value(cell_value):
#             continue
#         # Skip "Restated"
#         if isinstance(cell_value, str) and cell_value.strip().lower() == "restated":
#             continue
#         # Found a real value (including zero)
#         return True
#     return False

# # Function to check if column D has "Restated"
# def has_restated_in_d(sheet, row):
#     cell_d = sheet[f'D{row}']
#     if cell_d.value and isinstance(cell_d.value, str):
#         return cell_d.value.strip().lower() == "restated"
#     return False

# # Function to check if a row is hidden
# def is_row_hidden(sheet, row):
#     try:
#         return sheet.row_dimensions[row].hidden
#     except:
#         return False

# # Function to check if a column is completely blank
# def is_column_blank(sheet, col):
#     col_letter = get_column_letter(col)
#     for row in range(4, sheet.max_row + 1):  # Check from row 4 onwards
#         if is_row_hidden(sheet, row):
#             continue
#         cell_value = sheet[f'{col_letter}{row}'].value
#         if not is_blank_value(cell_value):
#             return False
#     return True

# # Function to check if a row should be ignored (Forecast-related or duplicate headers)
# def should_ignore_row(cell_a_value, header_texts):
#     if not cell_a_value:
#         return False
#     text = str(cell_a_value).strip().lower()
    
#     # Ignore rows starting with "forecast based on"
#     if text.startswith("forecast based on"):
#         return True
    
#     # Ignore rows that match header texts (duplicate headers in data)
#     if text in [h.lower() for h in header_texts if h]:
#         return True
    
#     return False

# # Function to check if header value is invalid (contains #REF!, #N/A, etc.)
# def is_invalid_header(value):
#     if not value:
#         return True
#     text = str(value).strip().lower()
#     # Check for Excel error values
#     if any(error in text for error in ['#ref!', '#n/a', '#value!', '#div/0!', '#name?', '#null!', '#num!']):
#         return True
#     return False

# # Function to make header names unique by appending suffix if duplicates
# def make_unique_headers(headers_dict):
#     seen = {}
#     unique_headers = {}
#     for col, header in headers_dict.items():
#         if header in seen:
#             seen[header] += 1
#             unique_headers[col] = f"{header}_{seen[header]}"
#         else:
#             seen[header] = 0
#             unique_headers[col] = header
#     return unique_headers

# # Function to format Excel output with styling
# def format_excel_output(df, company_info):
#     # Create a new workbook
#     output = io.BytesIO()
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Flattened Data"
    
#     # Define styles
#     dark_blue_fill = PatternFill(start_color="00003366", end_color="00003366", fill_type="solid")
#     white_font = Font(color="FFFFFF", bold=True)
#     wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
#     thin_border = Border(
#         left=Side(style='thin'),
#         right=Side(style='thin'),
#         top=Side(style='thin'),
#         bottom=Side(style='thin')
#     )
    
#     # Write column headers (starting from row 1) - Dark Blue with White Text
#     for col_idx, col_name in enumerate(df.columns, start=1):
#         # Replace first column header with company info
#         if col_idx == 1:
#             cell = ws.cell(row=1, column=col_idx, value=company_info)
#         else:
#             cell = ws.cell(row=1, column=col_idx, value=col_name)
#         cell.fill = dark_blue_fill
#         cell.font = white_font
#         cell.alignment = wrap_alignment
#         cell.border = thin_border
    
#     # Write data rows (starting from row 2)
#     for row_idx, (index_val, row_data) in enumerate(df.iterrows(), start=2):
#         # Write row header (vertical heading) with dark blue
#         cell = ws.cell(row=row_idx, column=1, value=index_val)
#         cell.fill = dark_blue_fill
#         cell.font = white_font
#         cell.alignment = wrap_alignment
#         cell.border = thin_border
        
#         # Write data values with word wrap and borders
#         for col_idx, (col_name, value) in enumerate(row_data.items(), start=2):
#             cell = ws.cell(row=row_idx, column=col_idx, value=value)
#             cell.alignment = wrap_alignment
#             cell.border = thin_border
    
#     # Adjust column widths
#     max_col = len(df.columns) + 1
#     for col_idx in range(1, max_col + 1):
#         ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
#     # Save workbook
#     wb.save(output)
#     output.seek(0)
#     return output

# # Function to process the sheet and flatten it dynamically
# def process_sheet(sheet):
#     # Step 1: Process A1-A3 to create company info string
#     a1_text = sheet['A1'].value or ""
#     a2_text = sheet['A2'].value or ""
#     a3_text = sheet['A3'].value or ""
#     # Combine with underscores for the new header format
#     company_info = f"{a1_text}_{a2_text}_{a3_text}".strip("_")
    
#     # Dynamic Header Construction: Scan for merged cells and build hierarchies
#     # Create a mapping of column index to header name
#     column_to_header = {}
#     processed_cols = set()
#     header_texts = []
    
#     # Multi-level header processing - handle 3 levels of hierarchy
#     for col in range(1, sheet.max_column + 1):
#         if col in processed_cols:
#             continue
            
#         col_letter = get_column_letter(col)
        
#         # Get values from all three header rows
#         row1_cell = sheet[f'{col_letter}1']
#         row2_cell = sheet[f'{col_letter}2']
#         row3_cell = sheet[f'{col_letter}3']
        
#         # Check if column is in a merged range
#         merged_range = get_merged_range(sheet, row1_cell)
        
#         if merged_range and merged_range.min_row == 1:
#             # Parent header from row 1
#             parent = str(row1_cell.value or "")
            
#             if is_invalid_header(parent) or parent.lower() == 'comments':
#                 processed_cols.add(col)
#                 continue
            
#             header_texts.append(parent.strip())
            
#             # Process each column in the merged range
#             for merged_col in range(merged_range.min_col, merged_range.max_col + 1):
#                 merged_col_letter = get_column_letter(merged_col)
                
#                 # Get child from row 2
#                 child_cell = sheet[f'{merged_col_letter}2']
#                 child_val = str(child_cell.value or "").strip()
                
#                 # Get grandchild from row 3
#                 grandchild_cell = sheet[f'{merged_col_letter}3']
#                 grandchild_val = str(grandchild_cell.value or "").strip()
                
#                 # Build header based on available levels
#                 if grandchild_val and not is_invalid_header(grandchild_val):
#                     if child_val and not is_invalid_header(child_val):
#                         # Three levels: Parent_Child_Grandchild
#                         column_to_header[merged_col] = f"{parent}_{child_val}_{grandchild_val}"
#                         header_texts.append(child_val)
#                         header_texts.append(grandchild_val)
#                     else:
#                         # Two levels: Parent_Grandchild
#                         column_to_header[merged_col] = f"{parent}_{grandchild_val}"
#                         header_texts.append(grandchild_val)
#                 elif child_val and not is_invalid_header(child_val):
#                     # Two levels: Parent_Child
#                     column_to_header[merged_col] = f"{parent}_{child_val}"
#                     header_texts.append(child_val)
#                 else:
#                     # One level: Parent only
#                     column_to_header[merged_col] = parent
                
#                 processed_cols.add(merged_col)
#         else:
#             # Single cell header (not part of a merged range in row 1)
#             if row1_cell.value and not is_invalid_header(row1_cell.value):
#                 column_to_header[col] = str(row1_cell.value)
#                 header_texts.append(str(row1_cell.value).strip())
#                 processed_cols.add(col)
    
#     # Set column A header
#     column_to_header[1] = company_info
    
#     # Make headers unique
#     column_to_header = make_unique_headers(column_to_header)
    
#     # Identify and mark blank columns for removal
#     blank_columns = set()
#     for col in range(1, sheet.max_column + 1):
#         if is_column_blank(sheet, col):
#             blank_columns.add(col)
    
#     # Now process rows dynamically for vertical headings and data
#     current_parent = None
#     current_child = None
#     rows_data = []
    
#     for row in range(4, sheet.max_row + 1):
#         # Skip hidden rows
#         if is_row_hidden(sheet, row):
#             continue
            
#         cell_a = sheet[f'A{row}']
#         cell_a_value = cell_a.value
        
#         # Skip forecast-related rows and duplicate headers
#         if should_ignore_row(cell_a_value, header_texts):
#             continue
        
#         if not cell_a_value:
#             continue
            
#         # Count leading spaces to determine hierarchy
#         cell_a_text = str(cell_a_value)
#         leading_spaces = count_leading_spaces(cell_a_text)
#         stripped_text = cell_a_text.strip()
        
#         # Check if rest of row has values
#         has_values = has_row_values(sheet, row, sheet.max_column)
        
#         # Check for parent row
#         is_parent = (
#             is_light_blue(cell_a) or 
#             (leading_spaces == 0 and not has_values) or 
#             (leading_spaces == 0 and has_restated_in_d(sheet, row))
#         )
        
#         # Build vertical heading
#         if is_parent:
#             current_parent = stripped_text
#             current_child = None
#             vertical_heading = current_parent
#         elif leading_spaces == 0 and has_values:
#             current_child = stripped_text
#             vertical_heading = f"{current_parent}_{current_child}"
#         elif leading_spaces >= 2 and has_values and current_child:
#             vertical_heading = f"{current_parent}_{current_child}_{stripped_text}"
#         else:
#             if current_parent:
#                 current_child = stripped_text
#                 vertical_heading = f"{current_parent}_{current_child}"
#             else:
#                 continue
        
#         # Extract data for this row - use actual column numbers
#         row_dict = {}
#         for col in range(1, sheet.max_column + 1):
#             if col not in blank_columns and col in column_to_header:
#                 col_letter = get_column_letter(col)
#                 row_dict[column_to_header[col]] = sheet[f'{col_letter}{row}'].value
        
#         rows_data.append({'vertical_heading': vertical_heading, 'data': row_dict})
    
#     # Create DataFrame with proper column alignment
#     if rows_data:
#         # Get all unique column headers in order
#         all_headers = [column_to_header[col] for col in sorted(column_to_header.keys()) 
#                       if col not in blank_columns]
        
#         # Build data rows
#         data_for_df = []
#         vertical_headings = []
#         for row_info in rows_data:
#             vertical_headings.append(row_info['vertical_heading'])
#             row_data = []
#             for header in all_headers:
#                 row_data.append(row_info['data'].get(header))
#             data_for_df.append(row_data)
        
#         df = pd.DataFrame(data_for_df, columns=all_headers, index=vertical_headings)
#     else:
#         df = pd.DataFrame()
    
#     return df, company_info

# # Streamlit App
# st.title("Excel Flattener App")
# st.write("Upload an Excel file, select a tab, and download the flattened data.")

# uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

# if uploaded_file is not None:
#     try:
#         # Load workbook with data_only=True to get computed values, not formulas
#         wb = openpyxl.load_workbook(uploaded_file, data_only=True)
#         sheet_names = wb.sheetnames
#         selected_sheet = st.selectbox("Select a tab", sheet_names)
        
#         if st.button("Process and Flatten"):
#             sheet = wb[selected_sheet]
#             flattened_df, company_info = process_sheet(sheet)
            
#             st.write("Flattened Data Preview:")
#             st.dataframe(flattened_df)
            
#             # Format and prepare Excel for download
#             formatted_excel = format_excel_output(flattened_df, company_info)
            
#             st.download_button(
#                 label="Download Formatted Excel",
#                 data=formatted_excel,
#                 file_name=f"{selected_sheet}_flattened_formatted.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
#     except Exception as e:
#         st.error(f"Error processing file: {str(e)}")


import streamlit as st
import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io

# Function to check if a cell has light blue background (adjust RGB if needed)
def is_light_blue(cell):
    try:
        return cell.fill.fgColor.rgb == 'FFD3D3D3'
    except:
        return False

# Function to get merged range for a cell
def get_merged_range(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range
    return None

# Function to check if a value is truly blank (None, empty string, or whitespace only)
def is_blank_value(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False

# Function to count leading spaces in a string
def count_leading_spaces(text):
    if not isinstance(text, str):
        return 0
    return len(text) - len(text.lstrip(' '))

# Function to check if rest of row has values (excluding "Restated")
def has_row_values(sheet, row, max_col):
    for col in range(2, max_col + 1):  # Start from B (column 2)
        col_letter = get_column_letter(col)
        cell_value = sheet[f'{col_letter}{row}'].value
        # Skip truly blank
        if is_blank_value(cell_value):
            continue
        # Skip "Restated"
        if isinstance(cell_value, str) and cell_value.strip().lower() == "restated":
            continue
        # Found a real value (including zero)
        return True
    return False

# Function to check if column D has "Restated"
def has_restated_in_d(sheet, row):
    cell_d = sheet[f'D{row}']
    if cell_d.value and isinstance(cell_d.value, str):
        return cell_d.value.strip().lower() == "restated"
    return False

# Function to check if a row is hidden
def is_row_hidden(sheet, row):
    try:
        return sheet.row_dimensions[row].hidden
    except:
        return False

# Function to check if a column is completely blank
def is_column_blank(sheet, col):
    col_letter = get_column_letter(col)
    for row in range(4, sheet.max_row + 1):  # Check from row 4 onwards
        if is_row_hidden(sheet, row):
            continue
        cell_value = sheet[f'{col_letter}{row}'].value
        if not is_blank_value(cell_value):
            return False
    return True

# Function to check if a row should be ignored (Forecast-related or duplicate headers)
def should_ignore_row(cell_a_value, header_texts):
    if not cell_a_value:
        return False
    text = str(cell_a_value).strip().lower()
    
    # Ignore rows starting with "forecast based on"
    if text.startswith("forecast based on"):
        return True
    
    # Ignore rows that match header texts (duplicate headers in data)
    if text in [h.lower() for h in header_texts if h]:
        return True
    
    return False

# Function to check if header value is invalid (contains #REF!, #N/A, etc.)
def is_invalid_header(value):
    if not value:
        return True
    text = str(value).strip().lower()
    # Check for Excel error values
    if any(error in text for error in ['#ref!', '#n/a', '#value!', '#div/0!', '#name?', '#null!', '#num!']):
        return True
    return False

# Function to check if a vertical heading contains % indicator
def is_percentage_row(vertical_heading):
    if not vertical_heading:
        return False
    text = str(vertical_heading).lower()
    # Check if heading contains common percentage indicators
    return '%' in text or 'percent' in text or 'pct' in text

# Function to convert value to percentage format if needed
def format_percentage(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Convert to percentage and round to 2 decimal places
        return f"{round(value * 100, 2)}%"
    return value

# Function to make header names unique by appending suffix if duplicates
def make_unique_headers(headers_dict):
    seen = {}
    unique_headers = {}
    for col, header in headers_dict.items():
        if header in seen:
            seen[header] += 1
            unique_headers[col] = f"{header}_{seen[header]}"
        else:
            seen[header] = 0
            unique_headers[col] = header
    return unique_headers

# Function to format Excel output with styling
def format_excel_output(df, company_info):
    # Create a new workbook
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flattened Data"
    
    # Define styles
    dark_blue_fill = PatternFill(start_color="00003366", end_color="00003366", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write column headers (starting from row 1) - Dark Blue with White Text
    for col_idx, col_name in enumerate(df.columns, start=1):
        # Replace first column header with company info
        if col_idx == 1:
            cell = ws.cell(row=1, column=col_idx, value=company_info)
        else:
            cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
    
    # Write data rows (starting from row 2)
    for row_idx, (index_val, row_data) in enumerate(df.iterrows(), start=2):
        # Write row header (vertical heading) with dark blue
        cell = ws.cell(row=row_idx, column=1, value=index_val)
        cell.fill = dark_blue_fill
        cell.font = white_font
        cell.alignment = wrap_alignment
        cell.border = thin_border
        
        # Check if this is a percentage row
        is_pct_row = is_percentage_row(index_val)
        
        # Write data values with word wrap and borders
        for col_idx, (col_name, value) in enumerate(row_data.items(), start=2):
            # Format as percentage if this is a percentage row
            if is_pct_row:
                formatted_value = format_percentage(value)
            else:
                formatted_value = value
                
            cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
            cell.alignment = wrap_alignment
            cell.border = thin_border
    
    # Adjust column widths
    max_col = len(df.columns) + 1
    for col_idx in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Save workbook
    wb.save(output)
    output.seek(0)
    return output

# Function to process the sheet and flatten it dynamically
def process_sheet(sheet):
    # Step 1: Process A1-A3 to create company info string
    a1_text = sheet['A1'].value or ""
    a2_text = sheet['A2'].value or ""
    a3_text = sheet['A3'].value or ""
    # Combine with underscores for the new header format
    company_info = f"{a1_text}_{a2_text}_{a3_text}".strip("_")
    
    # Dynamic Header Construction: Scan for merged cells and build hierarchies
    # Create a mapping of column index to header name
    column_to_header = {}
    processed_cols = set()
    header_texts = []
    
    # Multi-level header processing - handle 3 levels of hierarchy
    for col in range(1, sheet.max_column + 1):
        if col in processed_cols:
            continue
            
        col_letter = get_column_letter(col)
        
        # Get values from all three header rows
        row1_cell = sheet[f'{col_letter}1']
        row2_cell = sheet[f'{col_letter}2']
        row3_cell = sheet[f'{col_letter}3']
        
        # Check if column is in a merged range
        merged_range = get_merged_range(sheet, row1_cell)
        
        if merged_range and merged_range.min_row == 1:
            # Parent header from row 1
            parent = str(row1_cell.value or "")
            
            if is_invalid_header(parent) or parent.lower() == 'comments':
                processed_cols.add(col)
                continue
            
            header_texts.append(parent.strip())
            
            # Process each column in the merged range
            for merged_col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_col_letter = get_column_letter(merged_col)
                
                # Get child from row 2
                child_cell = sheet[f'{merged_col_letter}2']
                child_val = str(child_cell.value or "").strip()
                
                # Get grandchild from row 3
                grandchild_cell = sheet[f'{merged_col_letter}3']
                grandchild_val = str(grandchild_cell.value or "").strip()
                
                # Build header based on available levels
                if grandchild_val and not is_invalid_header(grandchild_val):
                    if child_val and not is_invalid_header(child_val):
                        # Three levels: Parent_Child_Grandchild
                        column_to_header[merged_col] = f"{parent}_{child_val}_{grandchild_val}"
                        header_texts.append(child_val)
                        header_texts.append(grandchild_val)
                    else:
                        # Two levels: Parent_Grandchild
                        column_to_header[merged_col] = f"{parent}_{grandchild_val}"
                        header_texts.append(grandchild_val)
                elif child_val and not is_invalid_header(child_val):
                    # Two levels: Parent_Child
                    column_to_header[merged_col] = f"{parent}_{child_val}"
                    header_texts.append(child_val)
                else:
                    # One level: Parent only
                    column_to_header[merged_col] = parent
                
                processed_cols.add(merged_col)
        else:
            # Single cell header (not part of a merged range in row 1)
            if row1_cell.value and not is_invalid_header(row1_cell.value):
                column_to_header[col] = str(row1_cell.value)
                header_texts.append(str(row1_cell.value).strip())
                processed_cols.add(col)
    
    # Set column A header
    column_to_header[1] = company_info
    
    # Make headers unique
    column_to_header = make_unique_headers(column_to_header)
    
    # Identify and mark blank columns for removal
    blank_columns = set()
    for col in range(1, sheet.max_column + 1):
        if is_column_blank(sheet, col):
            blank_columns.add(col)
    
    # Now process rows dynamically for vertical headings and data
    current_parent = None
    current_child = None
    rows_data = []
    
    for row in range(4, sheet.max_row + 1):
        # Skip hidden rows
        if is_row_hidden(sheet, row):
            continue
            
        cell_a = sheet[f'A{row}']
        cell_a_value = cell_a.value
        
        # Skip forecast-related rows and duplicate headers
        if should_ignore_row(cell_a_value, header_texts):
            continue
        
        if not cell_a_value:
            continue
            
        # Count leading spaces to determine hierarchy
        cell_a_text = str(cell_a_value)
        leading_spaces = count_leading_spaces(cell_a_text)
        stripped_text = cell_a_text.strip()
        
        # Check if rest of row has values
        has_values = has_row_values(sheet, row, sheet.max_column)
        
        # Check for parent row
        is_parent = (
            is_light_blue(cell_a) or 
            (leading_spaces == 0 and not has_values) or 
            (leading_spaces == 0 and has_restated_in_d(sheet, row))
        )
        
        # Build vertical heading
        if is_parent:
            current_parent = stripped_text
            current_child = None
            vertical_heading = current_parent
        elif leading_spaces == 0 and has_values:
            current_child = stripped_text
            vertical_heading = f"{current_parent}_{current_child}"
        elif leading_spaces >= 2 and has_values and current_child:
            vertical_heading = f"{current_parent}_{current_child}_{stripped_text}"
        else:
            if current_parent:
                current_child = stripped_text
                vertical_heading = f"{current_parent}_{current_child}"
            else:
                continue
        
        # Extract data for this row - use actual column numbers
        row_dict = {}
        for col in range(1, sheet.max_column + 1):
            if col not in blank_columns and col in column_to_header:
                col_letter = get_column_letter(col)
                row_dict[column_to_header[col]] = sheet[f'{col_letter}{row}'].value
        
        rows_data.append({'vertical_heading': vertical_heading, 'data': row_dict})
    
    # Create DataFrame with proper column alignment
    if rows_data:
        # Get all unique column headers in order
        all_headers = [column_to_header[col] for col in sorted(column_to_header.keys()) 
                      if col not in blank_columns]
        
        # Build data rows
        data_for_df = []
        vertical_headings = []
        for row_info in rows_data:
            vertical_headings.append(row_info['vertical_heading'])
            row_data = []
            for header in all_headers:
                row_data.append(row_info['data'].get(header))
            data_for_df.append(row_data)
        
        df = pd.DataFrame(data_for_df, columns=all_headers, index=vertical_headings)
    else:
        df = pd.DataFrame()
    
    return df, company_info

# Streamlit App
st.title("Excel Flattener App")
st.write("Upload an Excel file, select a tab, and download the flattened data.")

uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

if uploaded_file is not None:
    try:
        # Load workbook with data_only=True to get computed values, not formulas
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        sheet_names = wb.sheetnames
        selected_sheet = st.selectbox("Select a tab", sheet_names)
        
        if st.button("Process and Flatten"):
            sheet = wb[selected_sheet]
            flattened_df, company_info = process_sheet(sheet)
            
            st.write("Flattened Data Preview:")
            st.dataframe(flattened_df)
            
            # Format and prepare Excel for download
            formatted_excel = format_excel_output(flattened_df, company_info)
            
            st.download_button(
                label="Download Formatted Excel",
                data=formatted_excel,
                file_name=f"{selected_sheet}_flattened_formatted.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    except Exception as e:
        st.error(f"Error processing file: {str(e)}")
