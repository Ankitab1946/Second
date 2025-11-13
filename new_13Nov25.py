import streamlit as st
import pandas as pd
import openpyxl
import xlrd
import io
import tempfile

st.set_page_config(page_title="Financial Excel Flattener", layout="wide")


# ----------------------------------
# Convert XLS → XLSX automatically
# ----------------------------------
def convert_xls_to_xlsx(uploaded_file):
    try:
        # read xls using xlrd
        book = xlrd.open_workbook(file_contents=uploaded_file.read())
        sheet = book.sheet_by_index(0)

        # create new xlsx file in memory
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active

        for row in range(sheet.nrows):
            for col in range(sheet.ncols):
                ws.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

        wb.save(output)
        output.seek(0)
        return output
    except:
        return None


# ----------------------------------------------------
# Flatten Logic
# ----------------------------------------------------
def load_excel_flatten(file, sheet_name=0):
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

    max_row = ws.max_row
    max_col = ws.max_column

    grid = [[None for _ in range(max_col)] for _ in range(max_row)]

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            grid[r-1][c-1] = ws.cell(r, c).value

    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row_m, max_col_m = merged.bounds
        top_left_val = grid[min_row-1][min_col-1]
        for r in range(min_row, max_row_m + 1):
            for c in range(min_col, max_col_m + 1):
                grid[r-1][c-1] = top_left_val

    df = pd.DataFrame(grid)

    df = df.ffill(axis=0)
    df = df.ffill(axis=1)

    header_row = df.applymap(lambda x: isinstance(x, (int, float))).any(axis=1).idxmax()
    header_block = df.iloc[:header_row]

    flat_headers = (
        header_block.astype(str)
        .replace("nan", "")
        .apply(lambda col: [c for c in col if c != ""])
        .apply(lambda col: "_".join(col))
    )

    df.columns = flat_headers
    return df.iloc[header_row:].reset_index(drop=True)


# ----------------------------------------------------
# Streamlit App
# ----------------------------------------------------
st.title("📊 Financial Excel Flattener (XLS + XLSX Supported)")

uploaded = st.file_uploader("Upload Financial Excel", type=["xls", "xlsx"])

if uploaded:
    file_bytes = uploaded.read()
    uploaded.seek(0)

    if uploaded.name.endswith(".xls"):
        st.warning("Detected .xls file → converting to .xlsx automatically...")
        converted_file = convert_xls_to_xlsx(uploaded)
        if converted_file is None:
            st.error("Could not convert .xls file. It may be corrupted.")
        else:
            excel_file = converted_file
    else:
        excel_file = uploaded

    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet_names = wb.sheetnames

    sheet = st.selectbox("Select Sheet", sheet_names)

    if st.button("Flatten File"):
        df_flat = load_excel_flatten(excel_file, sheet)
        st.dataframe(df_flat, use_container_width=True)

        out = io.BytesIO()
        with pd.ExcelWriter(out, engine="xlsxwriter") as writer:
            df_flat.to_excel(writer, index=False)

        st.download_button(
            "📥 Download Flattened Excel",
            out.getvalue(),
            "flattened_financials.xlsx"
        )