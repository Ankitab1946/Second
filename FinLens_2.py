import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO

st.set_page_config(page_title="Financials Header & Row Flattener (Fixed)", layout="wide")

# -------------------------
# Helpers
# -------------------------
def safe_cell_value(ws, row, col):
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()

def is_fill_blue(cell):
    """
    Robust check for 'light blue' fills. openpyxl can represent colors in various ways.
    We'll check:
      - rgb string (like 'FFDBEEF3' or 'DBEEF3' or 'FF' prefixed)
      - indexed/theme: fallback to checking patternFill or presence of fill at all.
    This is heuristic — adjust hex palette if your blue shade differs.
    """
    try:
        fill = cell.fill
        if not fill:
            return False
        fg = fill.fgColor
        if not fg:
            return False
        # rgb property may be None
        rgb = getattr(fg, "rgb", None)
        if rgb:
            # Normalize: sometimes rgb like 'FFDBEEF3' or 'DBEEF3'
            rgb_norm = rgb.upper().lstrip("0X")
            # target suffixes we consider "blue-ish"
            blue_suffixes = ("DCE6F1", "DBEEF3", "C6D9F1", "EAF3FF", "DBECFF")
            for s in blue_suffixes:
                if rgb_norm.endswith(s):
                    return True
        # If indexed or theme color present, attempt to treat as non-blue unless explicit rgb matched
        return False
    except Exception:
        return False

# -------------------------
# Header extraction (3 rows)
# -------------------------
def extract_flattened_header(ws, header_rows=3):
    max_cols = ws.max_column

    # initialize header matrix: header_rows x max_cols
    header_matrix = [["" for _ in range(max_cols)] for _ in range(header_rows)]

    # read values into header_matrix
    for r in range(1, header_rows + 1):
        for c in range(1, max_cols + 1):
            header_matrix[r - 1][c - 1] = safe_cell_value(ws, r, c)

    # Propagate merged cell values INTO header_matrix but ONLY for rows within header_rows
    for merge_range in ws.merged_cells.ranges:
        # merge_range.min_row may be > header_rows -> skip
        if merge_range.max_row < 1 or merge_range.min_row > header_rows:
            continue
        val = safe_cell_value(ws, merge_range.min_row, merge_range.min_col)
        # clamp rows to header_rows
        r_start = max(merge_range.min_row, 1)
        r_end = min(merge_range.max_row, header_rows)
        for r in range(r_start, r_end + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                if 1 <= c <= max_cols:
                    header_matrix[r - 1][c - 1] = val

    # build flattened headers for each column
    final_headers = []
    for col in range(max_cols):
        h_parts = []
        h1 = header_matrix[0][col].strip()
        h2 = header_matrix[1][col].strip() if header_rows >= 2 else ""
        h3 = header_matrix[2][col].strip() if header_rows >= 3 else ""

        # Special: first column combine A1+A2+A3 into one header (as requested)
        if col == 0:
            combined = "_".join(p for p in (h1, h2, h3) if p).strip("_")
            final_headers.append(combined if combined else f"col_{col+1}")
            continue

        # Ignore 'Comments' as parent header
        if h1 and h1.lower() == "comments":
            # keep an empty identifiable header so column counts remain aligned
            final_headers.append("Comments")
            continue

        # general combine non-empty levels with underscore
        for part in (h1, h2, h3):
            if part:
                h_parts.append(part)
        header_name = "_".join(h_parts).strip("_")
        final_headers.append(header_name if header_name else f"col_{col+1}")

    return final_headers

# -------------------------
# Build dataframe rows explicitly from sheet (starting row 4)
# -------------------------
def build_dataframe_from_ws(ws, headers, start_row=4):
    max_cols = ws.max_column
    max_row = ws.max_row

    rows = []
    for r in range(start_row, max_row + 1):
        row_vals = []
        # read exactly max_cols cells to match header length
        for c in range(1, max_cols + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        rows.append(row_vals)

    df = pd.DataFrame(rows, columns=headers)
    # reset index for easier mapping with ws (df idx 0 corresponds to Excel row start_row)
    return df

# -------------------------
# Build vertical hierarchy labels
# -------------------------
def build_vertical_labels(df, ws, data_start_row=4):
    parent_for_index = {}
    last_parent = ""

    for idx in range(len(df)):
        excel_row = data_start_row + idx
        # first column text
        cell_text = safe_cell_value(ws, excel_row, 1)
        normalized = cell_text.strip()

        # ignore obvious note rows
        if normalized.lower() in ("", "forecast based on research"):
            parent_for_index[idx] = last_parent
            continue

        # detect blue fill on column A cell
        cell = ws.cell(row=excel_row, column=1)
        if is_fill_blue(cell):
            last_parent = normalized
            parent_for_index[idx] = last_parent
        else:
            parent_for_index[idx] = last_parent

    # now build final label for each row
    final_labels = []
    for idx in range(len(df)):
        parent = parent_for_index.get(idx, "")
        main = safe_cell_value(ws, data_start_row + idx, 1)

        if main.strip().lower() in ("", "forecast based on research"):
            final_labels.append("")  # keep blank for ignored rows
            continue

        # if row text looks like a "% change" variant -> try to attach to previous non-empty child
        low = main.lower()
        if "% change" in low or "%change" in low or low.strip().endswith("%"):
            # find previous non-empty child within df rows
            prev_child = ""
            # search backwards for the previous non-empty non-parent row label text
            for back in range(idx - 1, -1, -1):
                candidate = safe_cell_value(ws, data_start_row + back, 1)
                if candidate.strip() and candidate.strip().lower() not in ("forecast based on research", ""):
                    prev_child = candidate.strip()
                    break
            if parent:
                if prev_child:
                    final_labels.append(f"{parent}_{prev_child}_%Change")
                else:
                    final_labels.append(f"{parent}_%Change")
            else:
                # no parent found; just keep the child label
                if prev_child:
                    final_labels.append(f"{prev_child}_%Change")
                else:
                    final_labels.append("%Change")
        else:
            # normal parent_child
            if parent:
                final_labels.append(f"{parent}_{main.strip()}")
            else:
                final_labels.append(main.strip())

    return final_labels

# -------------------------
# Streamlit UI
# -------------------------
st.title("📊 Financials Header & Row Flattener — Fixed")

uploaded = st.file_uploader("Upload Financials.xlsx", type=["xlsx"])
if uploaded:
    try:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
    except Exception as e:
        st.error(f"Failed to open workbook: {e}")
        st.stop()

    sheets = wb.sheetnames
    sheet_name = st.selectbox("Select sheet", sheets)

    if sheet_name:
        ws = wb[sheet_name]
        st.success(f"Loaded sheet: {sheet_name} (rows: {ws.max_row}, cols: {ws.max_column})")

        # Extract headers (safe)
        headers = extract_flattened_header(ws, header_rows=3)
        st.subheader("Flattened headers (first 50 shown)")
        st.write(headers[:50])

        # Build dataframe explicitly from cells to ensure column count matches headers
        df = build_dataframe_from_ws(ws, headers, start_row=4)

        st.subheader("Raw extracted data (first 10 rows)")
        st.dataframe(df.head(10), use_container_width=True)

        # Vertical hierarchy
        final_labels = build_vertical_labels(df, ws, data_start_row=4)
        df.insert(0, "Final_Row_Label", final_labels)

        st.subheader("Final flattened output (first 20 rows)")
        st.dataframe(df.head(20), use_container_width=True)

        # Download
        towrite = BytesIO()
        with pd.ExcelWriter(towrite, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Flattened")
        towrite.seek(0)

        st.download_button(
            "⬇️ Download Flattened Excel",
            towrite,
            file_name="Flattened_Financials_fixed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
